import openpyxl
import collections
from openpyxl.styles.borders import Border, Side
import datetime


# 前提
# 1枚目のシートにはクーポン発行使用リスト
# 2枚目のシートにはスタンプカード積み立てリスト
# のエクセルファイルをresult.xlsxとしてこのpythonファイルと同じ階層に保存しておくこと。
# かつ、1,2枚目のシートにはA3からデータが入っていること。

# 使用するエクセルファイルを読み込み
wb = openpyxl.load_workbook('result.xlsx')

# 未加工データが入っているsheetの代入
# クーポン発行使用リスト
ori_coupon_sheet = wb.worksheets[0]
 # スタンプカード積み立てリスト
ori_stack_sheet = wb.worksheets[1]

# 羅線前情報
side = Side(style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)
# セルの色の前情報
fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='d3d3d3', bgColor='d3d3d3')


# ------------------------ここから処理開始-----------------------------

# ある月に来店した利用者のリストと来店回数-----------------------------

# customer_dataというシートを新規作成
wb.create_sheet('customer_data')
# 総シート数を取得し、データを追記していくcustomer_dataシートの番号を求め、sheet変数に代入
number_customer = len(wb.sheetnames)-1
sheet = wb.worksheets[number_customer]

row = 3
# 全利用者リスト
customer_lists =[]
# 全利用者来店履歴
all_customer_history =[]
for i in range(ori_stack_sheet.max_row-2):
    customer_history=ori_stack_sheet.cell(row=row,column=3).value
    all_customer_history.append(customer_history)
    row +=1
    if customer_history not in customer_lists:
        customer_lists.append(customer_history)
# 利用者別の来店回数カウント
# keyが顧客ID、valueが来店回数
visit_count={}
for i in customer_lists:
    total = all_customer_history.count(i)
    visit_count[i] = total
# 書き込み
sheet['A1'] = '顧客ID'
sheet['B1'] = '来店回数'
sheet['A1'].fill = fill
sheet['B1'].fill = fill


row = 2
for i in visit_count:
    sheet.cell(row = row , column = 1).value = i
    sheet.cell(row = row , column = 2).value = visit_count[i]
    row +=1
# 終了

# 来客頻度とその人数-------------------------------------------------
# customer_resultというシートを新規作成
wb.create_sheet('customer_result')
# 総シート数を取得し、データを追記していくcustomer_resultシートの番号を求め、sheet変数に代入
number_customer = len(wb.sheetnames)-1
sheet = wb.worksheets[number_customer]

times = []
for i in visit_count:
    times.append(visit_count[i])
# IDごとの回数が入ったリスト
each_times = collections.Counter(times)

sheet['A1']='来店回数'
sheet['B1']='人数'
sheet['A1'].fill = fill
sheet['B1'].fill = fill

row = 2
for i in each_times:
    sheet.cell(row = row,column = 1).value = i
    sheet.cell(row = row,column = 2).value = each_times[i]
    row +=1

# セルの整形（あってるか不明）
# for col in sheet.columns:
#     max_length = 0
#     column = col[0].column

#     for cell in col:
#         if len(str(cell.value)) > max_length:
#             max_length = len(str(cell.value))

#     adjusted_width = (max_length + 2) * 1.2
#     sheet.column_dimensions[column].width = adjusted_width
# 羅線作成
for row in sheet:
    for cell in row:
        sheet[cell.coordinate].border = border

# 終了


# クーポンの種類と使用率-----------------------------------------

# couponというシートを新規作成
wb.create_sheet('coupon')
# 総シート数を取得し、データを追記していくcouponシートの番号を求め、sheet変数に代入
number_coupon = len(wb.sheetnames)-1
sheet = wb.worksheets[number_coupon]
# 全クーポン種類
row = 3
coupon_lists =[]
all_coupon_history =[]
for i in range(ori_coupon_sheet.max_row-2):
    coupon_history=ori_coupon_sheet.cell(row=row,column=5).value
    all_coupon_history.append(coupon_history)
    row +=1
    if coupon_history not in coupon_lists:
        coupon_lists.append(coupon_history)
# クーポン発行回数
# keyがクーポン名、valueが発行回数
coupon = {}
for i in coupon_lists:
    total = all_coupon_history.count(i)
    coupon[i] = total
# 書き込み
sheet['A1']='クーポン名'
sheet['B1']='クーポン発行数'
sheet['C1']='割合（総発行数100）'
sheet['D1']='期限切れ発行枚数'
sheet['E1']='期限切れ使用済み枚数'
sheet['F1']='消化率'
sheet['A1'].fill = fill
sheet['B1'].fill = fill
sheet['C1'].fill = fill
sheet['D1'].fill = fill
sheet['E1'].fill = fill
sheet['F1'].fill = fill
# A1とB1書き込み
row = 2
for i in coupon:
    sheet.cell(row = row , column = 1).value = i
    sheet.cell(row = row , column = 2).value = coupon[i]
    row +=1
# C1書き込み
row = 2
total = len(all_coupon_history)
for i in range(len(coupon)):
    sheet.cell(row=row,column=3).value =str(round(int(sheet.cell(row=row,column=2).value)/total*100,1)) +'%'
    row +=1
# D1書き込み、E1書き込み
# 今日の日付
today = datetime.date.today()
# 使用期限日時の取得
row = 3
count = 0
limited_count = 0
used_limited_count = 0
for i in coupon_lists:
    for j in range(ori_coupon_sheet.max_row-2):
        date = ori_coupon_sheet.cell(row = row , column = 8).value
        limit_datetime = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
        limit_date = datetime.date(limit_datetime.year, limit_datetime.month, limit_datetime.day)
        if ori_coupon_sheet.cell(row = row , column = 5).value == i and limit_date < today:
            limited_count += 1
        if ori_coupon_sheet.cell(row = row , column = 5) == i and limit_date < today and ori_coupon_sheet.cell(row = row , column = 10).value:
            used_limited_count += 1
    number = coupon_lists.index(i)
    sheet.cell(row = number + 1,column = 4).value = limited_count
    sheet.cell(row = number + 1,column = 5).value = used_limited_count

    limited_count = 0
    used_limited_count = 0

# F1書き込み
row = 2
for i in range(sheet.max_row-1):
    if int(sheet.cell(row = row , column = 4).value) == 0:
        sheet.cell(row = row , column = 6).value = '0'
        row +=1
    else:
        sheet.cell(row = row , column = 6).value = str(round(int(sheet.cell(row = row , column = 5).value)/int(sheet.cell(row = row , column = 4).value)*100,1))+'%'
        row +=1

# セルの整形（あってるか不明）
# for col in sheet.columns:
#     max_length = 0
#     column = col[0].column

#     for cell in col:
#         if len(str(cell.value)) > max_length:
#             max_length = len(str(cell.value))

#     adjusted_width = (max_length + 2) * 1.2
#     sheet.column_dimensions[column].width = adjusted_width
# 羅線作成
for row in sheet:
    for cell in row:
        sheet[cell.coordinate].border = border
# 終了



# 保存
wb.save('result.xlsx')
