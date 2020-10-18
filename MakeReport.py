import openpyxl

# 前提
# 1枚目のシートにはクーポン発行使用リスト
# 2枚目のシートにはスタンプカード積み立てリスト
# のエクセルファイルをresult.xlsxとしてこのpythonファイルと同じ階層に保存しておくこと。
# かつ、1,2枚目のシートにはA3からデータが入っていること。

# 使用するエクセルファイルを読み込み
wb = openpyxl.load_workbook('result.xlsx')

# 未加工データが入っているsheetの代入
# クーポン発行使用リスト
ori_coupon_sheet = wb.worksheets[0]
 # スタンプカード積み立てリスト
ori_stack_sheet = wb.worksheets[1]



# ここから処理開始

# ある月に来店した利用者のリストと来店回数

# customerというシートを新規作成
wb.create_sheet('customer')
# 総シート数を取得し、データを追記していくcustomerシートの番号を求め、sheet変数に代入
number_customer = len(wb.sheetnames)-1
sheet = wb.worksheets[number_customer]

row = 3
# 全利用者リスト
customer_lists =[]
# 全利用者来店履歴
all_customer_history =[]
for i in range(ori_stack_sheet.max_row-2):
    customer_history=ori_stack_sheet.cell(row=row,column=2).value
    all_customer_history.append(customer_history)
    row +=1
    if customer_history not in customer_lists:
        customer_lists.append(customer_history)
# 利用者別の来店回数カウント
# keyが顧客ID、valueが来店回数
visit_count={}
for i in customer_lists:
    total = all_customer_history.count(i)
    visit_count[i] = total
# 書き込み
sheet['A1'] = '顧客ID'
sheet['B1'] = '来店回数'
row = 2
for i in visit_count:
    sheet.cell(row = row , column = 1).value = i
    sheet.cell(row = row , column = 2).value = visit_count[i]
    row +=1
# 終了


# クーポンの種類と使用率

# couponというシートを新規作成
wb.create_sheet('coupon')
# 総シート数を取得し、データを追記していくcouponシートの番号を求め、sheet変数に代入
number_coupon = len(wb.sheetnames)-1
sheet = wb.worksheets[number_coupon]
# 全クーポン種類
row = 3
coupon_lists =[]
all_coupon_history =[]
for i in range(ori_coupon_sheet.max_row-2):
    coupon_history=ori_coupon_sheet.cell(row=row,column=4).value
    all_coupon_history.append(coupon_history)
    start +=1
    if coupon_history not in coupon_lists:
        coupon_lists.append(coupon_history)
# クーポン発行回数
# keyがクーポン名、valueが発行回数
coupon = {}
for i in coupon_lists:
    total = all_coupon_history.count(i)
    coupon[i] = total
# 書き込み
sheet['A1']='クーポン名'
sheet['B1']='クーポン発行数'
sheet['C1']='割合（総発行数100）'
sheet['D1']='使用済み枚数'
sheet['E1']='消化率'
# A1とB1書き込み
row = 2
for i in coupon:
    sheet.cell(row = row , column = 1).value = i
    sheet.cell(row = row , column = 2).value = coupon[i]
    row +=1
# C1書き込み
row = 2
total = len(all_coupon_history)
for i in range(len(coupon)):
    sheet.cell(row=row,column=3).value =str(int(sheet.cell(row=row,column=2).value)/total*100) +'%'
    row +=1
# D1書き込み
row =3
# 済みだったらリストに入れる
used_list = []
for i in range(ori_coupon_sheet.max_row-2):
    # ''の中は使用済みを表すテキスト
    if ori_coupon_sheet.cell(row = row , column = 10).value = '':
        used_list.append(ori_coupon_sheet.cell(row = row , column = 4).value)
    row +=1
# 使用済みの回数を計算し、書き込む
row = 2
for i in coupon:
    total = used_list.count(i)
    sheet.cell(row = row , column = 4).value = total
    row +=1
# E1書き込み
row = 2
for i in range(sheet.max_row-1):
    sheet.cell(row = row , column = 5).value = str(int(sheet.cell(row = row , column = 4).value)/int(sheet.cell(row = row , column = 2).value)*100)+'%'

# 終了
# 保存
wb.save('result.xlsx')
